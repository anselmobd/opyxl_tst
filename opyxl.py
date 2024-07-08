from openpyxl import load_workbook, Workbook, styles
from pprint import pprint


class EquiqXlsx:

    # columns read
    DESCR = 'Descr. Sint.'
    DATA = 'Dt.Aquisicao'
    QUANT = 'Quantidade'
    TIPO = 'Tipo Ativo'

    def __init__(self, filename, destino):
        self.filename = filename
        self.destino = destino

        self.origin_columns_idx = {
            self.DESCR: None,
            self.DATA: None,
            self.QUANT: None,
            self.TIPO: None,
        }
        self.destination = {}
        self.tags = []
        self.anos = []
        self.info_recipe = {
            'key':(
                {
                    'value': self.DESCR,
                    'transform': self.get_tag_descricao
                },
                {
                    'value': self.DATA,
                    'transform': self.get_ano_data
                },
            ),
            'value': {
                'value': self.QUANT,
            },
            'apply': self.aualiza_valor,
        }
        self.tags = {
            'MATRICIAL': 'matricial',
            'SERVIDOR': 'servidor',
            'MICROCOMPUTADOR': 'desktop',
            'DESKTOP': 'desktop',
            'COMPUTADOR': 'desktop',
            'PROJETOR': 'projetor',
            'MONITOR': 'monitor',
            'NOTEBOOK': 'notebook',
            'ULTRABOOK': 'notebook',
            'MACBOOK': 'notebook',
            'APPLE MAC': 'notebook',
            'ETIQUETA': 'termica',
            'SCANNER': 'scanner',
            'NOBREAK': 'nobreak',
            'FIREWALL': 'firewall',
            'CPAC': 'firewall',
            'SECURITY': 'firewall',
            'SANDBLAST': 'firewall',
            'ACESS PONTS': 'ap',
            'NCOMPUTING': 'ncomputing',
            'STORAGE': 'storage',
            'BACKUP': 'storage',
            'SWITCH': 'switch',
            'SWICH': 'switch',
            'TRANSCEIVER': 'transceiver',
            'IPAD': 'tablet',
            'TABLET': 'tablet',
            'CAMERA': 'camera',
            None: 'outros',
        }
        self.max_row=None

    def get_tag_descricao(self, value):
        for search, tag in self.tags.items():
            if search and value.find(search) > -1:
                return tag
        return self.tags[None]

    def get_ano_data(self, value):
        return value.year

    def aualiza_valor(self, key, value):
        try:
            old_value = self.destination[key]
        except KeyError as _:
            old_value = 0
        self.destination[key] = old_value + value

    def mount_info(self, recipe, row):
        if isinstance(recipe, tuple):
            step_info = []
            for step in recipe:
                step_info.append(self.mount_info(step, row))
            return tuple(step_info)
        else:
            value = row[recipe['value']]
            if 'transform' in recipe:
                value = recipe['transform'](value)
            return value

    def apply_destination_info(self, row):
        key = self.mount_info(self.info_recipe['key'], row)
        value = self.mount_info(self.info_recipe['value'], row)
        self.info_recipe['apply'](key, value)

    def load(self):
        self.wb = load_workbook(filename=self.filename)
        self.ws = self.wb.active

    def get_col_id(self, row):
        for idx_col, cell in enumerate(row):
            if cell.value in self.origin_columns_idx:
                self.origin_columns_idx[cell.value] = idx_col

    def data_append_item(self, idx_row, row):
        item = {}
        for name, idx_col in self.origin_columns_idx.items():
            item[name] = row[idx_col].value
        self.apply_destination_info(item)

    def item_valido(self, row):
        return not row[self.origin_columns_idx['Tipo Ativo']].value

    def walk_through(self):
        for idx_row, row in enumerate(
                self.ws.iter_rows(max_row=self.max_row), start=1):
            if idx_row == 1:
                self.get_col_id(row)
            else:
                if self.item_valido(row):
                    self.data_append_item(idx_row, row)

    def print(self):
        # pprint(self.destination)
        print(sum(self.destination.values()), 'itens')
        pprint(self.tags)
        pprint(self.anos)

    def create_xlsx(self):
        self.dest_wb = Workbook()
        self.dest_ws = self.dest_wb.active
        self.bold_font = styles.Font(bold=True)

    def get_tags_anos(self):
        tags = set()
        anos = set()
        for tag, ano in self.destination:
            tags.add(tag)
            anos.add(ano)
        self.tags = sorted(list(tags))
        self.anos = sorted(list(anos))

    def set_tags(self):
        tags = ['Tag \ Ano'] + self.tags + ['Totais']
        max_length = 0
        for line, tag in enumerate(tags, start=1):
            max_length = max(len(tag), max_length)
            cell = self.dest_ws.cell(row=line, column=1, value=tag)
            cell.font = self.bold_font
            cell.alignment = styles.Alignment(horizontal='center')  
        self.dest_ws.column_dimensions['A'].width = max_length + 2

    def set_anos(self):
        for coluna, ano in enumerate(self.anos + ['Totais'], start=2):
            cell = self.dest_ws.cell(row=1, column=coluna, value=ano)
            cell.font = self.bold_font
            self.dest_ws.column_dimensions[cell.column_letter].width = 6
            cell.alignment = styles.Alignment(horizontal='right')  
        # a última coluna, de totais, redefina largura para 8
        self.dest_ws.column_dimensions[cell.column_letter].width = 8

    def set_valores(self):
        for key, value in self.destination.items():
            tag, ano = key
            idx_tag = self.tags.index(tag) + 2
            idx_ano = self.anos.index(ano) + 2
            self.dest_ws.cell(row=idx_tag, column=idx_ano, value=value)

    def tot_anos(self):
        for coluna, _ in enumerate(self.anos, start=2):
            linha_totais = len(self.tags) + 2
            cell = self.dest_ws.cell(row=linha_totais, column=coluna)
            cell.font = self.bold_font
            letra = cell.column_letter
            cell.value = f'=SUM({letra}2:{letra}{linha_totais-1}'

    def tot_tags(self):
        # soma inclusive a linha de totais por ano
        for linha in range(2, len(self.tags)+3):
            coluna_totais = len(self.anos) + 2
            cell = self.dest_ws.cell(row=linha, column=coluna_totais)
            cell_esquerda = self.dest_ws.cell(row=linha, column=coluna_totais-1)
            letra_esquerda = cell_esquerda.column_letter
            cell.value = f'=SUM(B{linha}:{letra_esquerda}{linha}'
            cell.font = self.bold_font

    def make_analise(self):
        self.dest_ws['A1'] = 'Tag \ Ano'
        self.set_tags()
        self.set_anos()
        self.set_valores()
        self.tot_anos()
        self.tot_tags()

    def save(self):
        self.get_tags_anos()
        self.create_xlsx()
        self.make_analise()
        self.dest_wb.save(self.destino)

    def process(self):
        self.load()
        self.walk_through()
        self.save()
        # self.print()


if __name__ == '__main__':
    EquiqXlsx('equipamentos.xlsx', 'analise.xlsx').process()
